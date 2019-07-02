import openpyxl
import datetime
import pickle

CITY_NUM = 656 # the number of all the citis
DAY_SEC = 86400 # the seconds of one day
HUB_DISCOUNT = 0.45 # the discount a hub can take
MONEYCOST_RATIO = 1
TIMECOST_RATIO = 500 # (will be divided)


class Edge(object):
    start = 0
    end = 0
    distance = 0.0
    speed = 0.0
    delayTime = 0.0
    departureTime = 0.0
    unitCost = 0.0
    way = ''

    def __init__(self, start, end, distance, speed, delayTime, departureTime, unitCost):
        self.start = start
        self.end = end
        self.distance = distance
        self.speed = speed
        self.delayTime = delayTime
        self.departureTime = departureTime
        self.unitCost = unitCost

    def __init__(self):
        self.start = 0
        self.end = 0
        self.distance = 0.0
        self.speed = 0
        self.delayTime = 0.0
        self.departureTime = 0.0
        self.unitCost = 0.0

    def weight(self, order, arrivalTime): # order as argument for convenience
        waitTime = (self.departureTime - arrivalTime if (arrivalTime <= self.departureTime) else DAY_SEC - arrivalTime + self.departureTime) # the time to wait for departure
        if (order.isEmergency == 1):
            return self.distance / self.speed * 3600 + self.delayTime * 60 + waitTime / TIMECOST_RATIO
        else:
            return (self.distance / self.speed * 3600 + self.delayTime * 60 + waitTime) / TIMECOST_RATIO + (order.totalWeight * self.unitCost * self.distance / 50) / MONEYCOST_RATIO
        # return self.distance / self.speed * 3600 + self.delayTime * 60 + waitTime # only consider the time

    def hub_weight(self, order, arrivalTime):
        waitTime = (self.departureTime - arrivalTime if (arrivalTime <= self.departureTime) else DAY_SEC - arrivalTime + self.departureTime)  # the time to wait for departure
        if (order.isEmergency == 1):
            return self.distance / self.speed * 3600 + self.delayTime * 60 + waitTime /TIMECOST_RATIO
        else:
            return (self.distance / self.speed * 3600 + self.delayTime * 60 + waitTime) / TIMECOST_RATIO + (order.totalWeight * self.unitCost * self.distance / 50) * HUB_DISCOUNT /MONEYCOST_RATIO

    def waitTimeAndTransTime(self, arrivalTime):
        waitTime = (self.departureTime - arrivalTime if (arrivalTime <= self.departureTime) else DAY_SEC - arrivalTime + self.departureTime)  # the time to wait for departure
        return self.distance / self.speed * 3600 + self.delayTime * 60 + waitTime
class Vertex(object):
    index = 0
    edges = []

    def __init__(self):
        self.index = 0
        self.edges = []

    def __init__(self, index):
        self.index = index
        self.edges = []


class Order(object):
    index = 0
    start = 0
    end = 0
    orderTime = 0.0
    goods = 0
    amount = 0
    isEmergency = 0
    totalWeight = 0.0
    totalTime = 0

    def __init__(self, start, end, goods, amount, isEmergency):
        self.start = start
        self.end = end
        self.goods = goods
        self.amount = amount
        self.isEmergency = isEmergency

    def __init__(self):
        self.start = 0
        self.end = 0
        self.orderTime = 0.0
        self.goods = 0
        self.amount = 0
        self.isEmergency = 0

def makeGraph():
    graph = []
    for i in range(CITY_NUM):
        vertex = Vertex(i + 1)
        graph.append(vertex)

    wb = openpyxl.load_workbook('TableD-TransportationTools.xlsx')
    wb_distance = openpyxl.load_workbook('TableC-DistanceMatrix.xlsx')
    sheet1 = wb.get_sheet_by_name('Plane')
    sheet2 = wb.get_sheet_by_name('Ship')
    sheet3 = wb.get_sheet_by_name('Train')
    sheet4 = wb.get_sheet_by_name('Truck')
    sheet_distance = wb_distance.get_sheet_by_name('Sheet1')
    print('Open tables Done!')

    for i in range(2, sheet1.max_row + 1):
        departureCity = sheet1.cell(row=i, column=1).value
        arrivalCity = sheet1.cell(row=i, column=2).value
        delayTime = sheet1.cell(row=i, column=3).value
        speed = sheet1.cell(row=i, column=4).value
        unitCost = sheet1.cell(row=i, column=5).value
        departureTime = sheet1.cell(row=i, column=6).value
        edge = Edge()
        edge.start = departureCity
        edge.end = arrivalCity
        edge.speed = speed
        edge.distance = sheet_distance.cell(row=edge.start, column=edge.end).value
        edge.delayTime = delayTime
        edge.departureTime = departureTime.hour * 3600 + departureTime.minute * 60 + departureTime.second
        edge.unitCost = unitCost
        edge.way = 'Plane'
        graph[departureCity - 1].edges.append(edge)

    for i in range(2, sheet2.max_row + 1):
        departureCity = sheet2.cell(row=i, column=1).value
        arrivalCity = sheet2.cell(row=i, column=2).value
        delayTime = sheet2.cell(row=i, column=3).value
        speed = sheet2.cell(row=i, column=4).value
        unitCost = sheet2.cell(row=i, column=5).value
        departureTime = sheet2.cell(row=i, column=6).value
        edge = Edge()
        edge.start = departureCity
        edge.end = arrivalCity
        edge.speed = speed
        edge.distance = sheet_distance.cell(row=edge.start, column=edge.end).value
        edge.delayTime = delayTime
        edge.departureTime = departureTime.hour * 3600 + departureTime.minute * 60 + departureTime.second
        edge.unitCost = unitCost
        edge.way = 'Ship'
        graph[departureCity - 1].edges.append(edge)

    for i in range(2, sheet3.max_row + 1):
        departureCity = sheet3.cell(row=i, column=1).value
        arrivalCity = sheet3.cell(row=i, column=2).value
        delayTime = sheet3.cell(row=i, column=3).value
        speed = sheet3.cell(row=i, column=4).value
        unitCost = sheet3.cell(row=i, column=5).value
        departureTime = sheet3.cell(row=i, column=6).value
        edge = Edge()
        edge.start = departureCity
        edge.end = arrivalCity
        edge.speed = speed
        edge.distance = sheet_distance.cell(row=edge.start, column=edge.end).value
        edge.delayTime = delayTime
        edge.departureTime = departureTime.hour * 3600 + departureTime.minute * 60 + departureTime.second
        edge.unitCost = unitCost
        edge.way = 'Train'
        graph[departureCity - 1].edges.append(edge)

    for i in range(2, sheet4.max_row + 1):
        departureCity = sheet4.cell(row=i, column=1).value
        arrivalCity = sheet4.cell(row=i, column=2).value
        delayTime = sheet4.cell(row=i, column=3).value
        speed = sheet4.cell(row=i, column=4).value
        unitCost = sheet4.cell(row=i, column=5).value
        departureTime = sheet4.cell(row=i, column=6).value
        edge = Edge()
        edge.start = departureCity
        edge.end = arrivalCity
        edge.speed = speed
        edge.distance = sheet_distance.cell(row=edge.start, column=edge.end).value
        edge.delayTime = delayTime
        edge.departureTime = departureTime.hour * 3600 + departureTime.minute * 60 + departureTime.second
        edge.unitCost = unitCost
        edge.way = 'Truck'
        graph[departureCity - 1].edges.append(edge)

    return graph

if __name__ == '__main__':
    graph = makeGraph()
    output = open('graph.pkl', 'wb', -1)

    pickle.dump(graph, output)
    output.close()

    # open tables for others
    table_orders = openpyxl.load_workbook('TableA-Orders.xlsx')
    table_commodities = openpyxl.load_workbook('TableB-Commodities.xlsx')
    sheet_orders = table_orders.get_sheet_by_name('Sheet1')
    sheet_commodities = table_commodities.get_sheet_by_name('Sheet1')

    output_orders = open('sheet_orders.pkl', 'wb')
    output_commodities = open('sheet_commodities.pkl', 'wb')
    pickle.dump(sheet_orders, output_orders)
    pickle.dump(sheet_commodities, output_commodities)
    output_orders.close()
    output_commodities.close()

